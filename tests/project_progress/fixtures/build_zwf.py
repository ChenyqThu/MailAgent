"""从 v2 (4-sheet) 真实 xlsx 生成 ENBU 子集 fixture.

每个 sheet 抽 ENBU 前 8 + 非 ENBU 前 4 行 (验证 BU 过滤), 保留双行表头结构.
不应提交包含真实业务数据的 xlsx (.gitignore 已屏蔽 sample_*.xlsx).

Usage:
    V2_SRC_XLSX=/path/to/v2-source.xlsx python tests/project_progress/fixtures/build_zwf.py
"""

import os
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(ROOT))

DEFAULT_SRC = ROOT / "tmp" / "v2-source.xlsx"
SRC_XLSX = Path(os.environ.get("V2_SRC_XLSX", str(DEFAULT_SRC)))
OUT = Path(__file__).parent / "sample_zwf_4sheet.xlsx"


SHEETS_OF_INTEREST = [
    "Project  Ongoing",
    "2026-Project Shipped",
    "Project Suspended",
]
GUIDE_SHEET = "Filling-in & Reading Guide"

ENBU_TAKE = 8
NON_ENBU_TAKE = 4


def _read_sheet_raw(src: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(src, sheet_name=sheet_name, header=None, engine="openpyxl")


def _find_header_row(df: pd.DataFrame) -> Optional[int]:
    """找首个含 'BU' + 'Project Name' 的行索引 (header_idx)."""
    for i in range(min(5, len(df))):
        cells = [str(v).strip() if v is not None else "" for v in df.iloc[i].tolist()]
        if "BU" in cells and "Project Name" in cells:
            return i
    return None


def _row_has_bu(row, bu_col_idx: int, target: str) -> bool:
    val = row.iloc[bu_col_idx]
    return val is not None and str(val).strip() == target


def _select_subset(df: pd.DataFrame, header_idx: int) -> pd.DataFrame:
    """保留 row 0..header_idx (banner + header + 中文标签若有), 然后 8 ENBU + 4 非 ENBU."""
    headers = [str(v).strip() if v is not None else "" for v in df.iloc[header_idx].tolist()]
    bu_col_idx = headers.index("BU")

    # 检测中文标签行: header 下一行
    next_idx = header_idx + 1
    keep_top = header_idx + 1  # row 0..header_idx 全留
    if next_idx < len(df):
        next_cells = [str(v).strip() if v is not None else "" for v in df.iloc[next_idx].tolist()]
        if next_cells:
            first_non_empty = next((c for c in next_cells if c), "")
            zh_markers = ("事业部", "课组", "研发", "项目管理", "产品负责人")
            if any(m in first_non_empty for m in zh_markers):
                keep_top += 1  # 中文标签行也留

    top_rows = df.iloc[:keep_top]
    data = df.iloc[keep_top:]

    enbu_rows = data[data.iloc[:, bu_col_idx].astype(str).str.strip() == "TPS-ENBU"].head(ENBU_TAKE)
    non_enbu_rows = data[data.iloc[:, bu_col_idx].astype(str).str.strip() != "TPS-ENBU"].head(NON_ENBU_TAKE)

    return pd.concat([top_rows, enbu_rows, non_enbu_rows], ignore_index=True)


def _maybe_read_guide(src: Path) -> Optional[pd.DataFrame]:
    """读 Filling-in Guide sheet (内容简短, 全部保留)."""
    try:
        return pd.read_excel(src, sheet_name=GUIDE_SHEET, header=None, engine="openpyxl")
    except Exception:
        return None


def main() -> int:
    if not SRC_XLSX.exists():
        print(f"ERROR: source xlsx not found at {SRC_XLSX}")
        print("Set SRC_XLSX env var to point to a real zwf xlsx.")
        return 1

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in SHEETS_OF_INTEREST:
        try:
            df = _read_sheet_raw(SRC_XLSX, sheet)
        except Exception as e:
            print(f"  skip {sheet!r}: {e}")
            continue
        header_idx = _find_header_row(df)
        if header_idx is None:
            print(f"  skip {sheet!r}: no BU+Project Name header found")
            continue
        subset = _select_subset(df, header_idx)
        ws = wb.create_sheet(sheet)
        for r_idx, row in enumerate(subset.itertuples(index=False), 1):
            for c_idx, val in enumerate(row, 1):
                # NaN → 空
                if isinstance(val, float) and val != val:
                    continue
                ws.cell(row=r_idx, column=c_idx, value=val)
        print(f"  wrote sheet {sheet!r} with {len(subset)} rows")

    # Guide sheet
    guide_df = _maybe_read_guide(SRC_XLSX)
    if guide_df is not None:
        ws = wb.create_sheet(GUIDE_SHEET)
        # 取前 10 行 (够测试 sheet 跳过逻辑)
        for r_idx, row in enumerate(guide_df.head(10).itertuples(index=False), 1):
            for c_idx, val in enumerate(row, 1):
                if isinstance(val, float) and val != val:
                    continue
                ws.cell(row=r_idx, column=c_idx, value=val)
        print(f"  wrote guide sheet with {min(10, len(guide_df))} rows")

    wb.save(OUT)
    print(f"\nWrote {OUT} ({OUT.stat().st_size:,} bytes)")

    # 验证读取
    check_wb = load_workbook(OUT)
    print(f"sheets: {check_wb.sheetnames}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
